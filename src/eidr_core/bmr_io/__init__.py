"""BMR workbook-surgery primitives — THE shared implementation.

Extracted 2026-08-03 (register Phase 3 item 1 / OVERLAPS rows 1-2) from
eidr-wikidata ``bmr/writer.py`` (canonical), unifying the audit copy in
BMR-Review that had drifted from it. Six primitives, canonical semantics:

* ``read_headers(ws)`` — header-row map {col_index: header_text}.
* ``count_family(headers, primary)`` — how many numbered copies of a family
  the template carries (returns the MAX index — robust to gaps, unlike the
  old audit-copy count-of-matches).
* ``rightmost_in(headers, members)`` — rightmost column of a family.
* ``expand_family(ws, primary, anchor_members, insert_members, want, headers)``
  — add family groups. The anchor/insert split is the operator's 2026-05-09
  direction (anchor = existing group-1 columns used to find the insertion
  point; insert = which columns new groups actually add). The pre-split
  behavior is the special case ``anchor_members=members, insert_members=members``
  (how the BMR-Review audit writer calls it).
* ``transplant(template_xlsx, edited_xlsx)`` — swap edited worksheet XML +
  shared strings into the template, preserving formulas/validations/macros.
* ``fix_shared_strings(xlsx_path)`` — inlineStr -> shared-string conversion
  for EPPlus-based BMR tooling compatibility.

Orchestration stays per-consumer BY DESIGN: eidr-wikidata's typed
``BMRWriter``/``FAMILIES`` and BMR-Review's dict-based multi-template
``write_bmr_files`` serve different jobs; only the workbook surgery was
duplicated.

READER HALF (added 2026-08-06, register R3 tail):

* ``open_sheet(path, sheet_name, ...)`` — STREAMING context manager
  yielding ``(headers, rows)``, where rows is a lazy iterator of
  ``(row_number, {header: value})``. Contributed by BMRtoAltID 2026-08-26
  for sheets too large to materialize (60k rows x 500+ columns). Emits
  absolute sheet row numbers, which cannot be reconstructed downstream
  once ``stop=None`` has skipped blanks and which any consumer reporting
  a defect to a human needs.
* ``read_sheet(path, sheet_name, ...)`` — the materializing form: a thin
  ``list()`` over ``open_sheet``, so both share ONE implementation of the
  header policy and the stop rules. Header-row map + one
  ``{header: value}`` dict per data row. Canonical semantics =
  eidr-wikidata ``scripts/combine_bmr_sheets.py::_read_chunk`` (the only
  reader that was already built on this module's constants). The two other
  portfolio readers stop early instead of reading everything, so the stop
  rule is a parameter — ``stop=None`` (read all, skip blank rows; combine),
  ``stop="blank_first_col"`` (halt when column A is empty; BMRtoAltID and
  the real BMR tool's behavior), ``stop="blank_row"`` (halt at the first
  fully-blank row; XML_to_JSON's BMR codec).
* ``family_layout(header_names, members)`` — SPARSE, index-preserving map
  ``{group_index: {member: actual_header}}`` for a repeating column family.
  Sparse-by-design: XML_to_JSON's ``_collect_numbered`` densifies groups
  before pairing ``Alt Title Class N``, which mis-aligns class/language
  when group 1 is blank — the shared layout keeps original indices so
  consumers can't repeat that bug. A bare member name (no number) counts
  as group 1, matching ``count_family``; a numbered ``"X 1"`` wins over a
  bare ``"X"`` if both somehow coexist.
* ``RepeatPlan`` — union-max repeat-group counts across records, with
  per-family minimums (folded in from eidr-dq ``flatten.py``, which
  hardcoded Director>=2 / Actor>=4 inside ``finalize``). The same
  union-max model backs combine's chunk merging and the writer's
  ``_max_counts``.
* ``pad_groups(items, count, width)`` — flatten one family's per-record
  values to exactly ``count`` groups of ``width`` cells, padding with
  ``fill``. Replaces the ~18 hand-written pad loops in flatten.py.

Consumer POLICY stays out of the reader by design: sheet auto-detection
strategies (XML_to_JSON's exact-tab allowlist vs BMRtoAltID's reserved-name
exclusion), value trimming/blank sentinels, the "." -> Proprietary domain
split, ShortDOI filtering, and IMDb/ISAN/V-ISAN singleton promotion all
differ per consumer and belong above this layer.
"""
from __future__ import annotations

import contextlib
import html
import logging
import os
import re
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Iterable, Iterator, Mapping, Sequence

log = logging.getLogger(__name__)

# Row 3 carries the column headers in every EIDR BMR template
# (identical in both pre-extraction writers).
HEADER_ROW = 3

# First data row in every EIDR BMR template (row 1-2 = banner, 3 = headers).
DATA_START = 4

__all__ = ["HEADER_ROW", "DATA_START", "read_headers", "count_family",
           "rightmost_in", "expand_family", "transplant",
           "fix_shared_strings",
           "open_sheet", "read_sheet", "family_layout", "RepeatPlan",
           "pad_groups"]


def _header_map(values: Iterable) -> dict[int, str]:
    """THE header-row policy, in one place: ``{1-based column: header}``.

    **CONTRACT — the key is an INTEGER column index, never a column
    letter.** Consumers order the inventory with
    ``[m[c] for c in sorted(m)]`` to recover sheet order; that identity
    holds only for integer keys. Letter keys would sort ``'AA'`` before
    ``'B'``, silently scrambling every column past Z — and stock BMR
    templates run to 602 columns, so this is not a hypothetical edge.
    Raised by XML_to_JSON 2026-08-25 (both its call sites depend on it);
    pinned by ``test_header_order_is_sheet_order_past_column_z``. Do not
    "modernize" this to openpyxl's own cell addressing.

    Blank and whitespace-only cells are skipped but do NOT stop the scan,
    and surviving columns keep their TRUE indices — a spacer must never
    hide, or renumber, the columns to its right. That single rule is the
    whole reason the shared reader exists (register R3: XML_to_JSON's two
    hand-rolled scans each broke at the first blank cell).

    Takes raw values rather than a worksheet because the two callers
    reach the row differently for a real reason — ``read_headers`` random-
    accesses a live worksheet, ``read_sheet`` streams in read-only mode
    where ``ws.cell`` is O(N) per access. Only the ACCESS differs; the
    policy is identical, so it lives here rather than in both.
    """
    out: dict[int, str] = {}
    for col_idx, v in enumerate(values, 1):
        if v is None:
            continue
        s = str(v).strip()
        if s:
            out[col_idx] = s
    return out


def read_headers(ws, header_row: int = HEADER_ROW) -> dict[int, str]:
    """Header map ``{1-based INTEGER column: header text}`` from a live
    worksheet. The integer key is a CONTRACT consumers sort on to recover
    sheet order — see ``_header_map``.

    ``header_row`` defaults to the template's row 3. It is a parameter
    because a consumer may DISCOVER the header row instead of assuming it
    — BMR-Review reads workbooks that have been round-tripped through
    review and locates row 3 by searching for ``Unique Row ID``. Without
    this argument that consumer could not adopt the shared reader at all,
    which is how it came to keep its own copy (register R3, 2026-08-25).

    For a path rather than an open worksheet, use ``read_sheet``, which
    also returns the data rows and streams instead of random-accessing.
    """
    return _header_map(
        ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)
    )

def count_family(headers: dict[int, str], primary: str) -> int:
    best = 0
    for h in headers.values():
        if h == primary:
            best = max(best, 1)
        elif h.startswith(primary + " "):
            # Non-numeric suffix means "not a numbered copy of this family",
            # not an error — skip it.
            with contextlib.suppress(ValueError):
                best = max(best, int(h[len(primary) + 1:]))
    return best

def rightmost_in(headers: dict[int, str], members: list[str]) -> int | None:
    right: int | None = None
    for col, h in headers.items():
        for m in members:
            if h == m or h.startswith(m + " "):
                right = max(right, col) if right is not None else col
    return right

def expand_family(ws, primary: str,
                   anchor_members: list[str], insert_members: list[str],
                   want: int, headers: dict[int, str]) -> None:
    """Add ``want - have`` new groups for the given family.

    ``anchor_members`` drives where new groups land (rightmost match
    in template). ``insert_members`` drives what columns are added
    per group; on each iteration the anchor advances by
    ``len(insert_members)`` to keep new groups packed together.
    """
    have = count_family(headers, primary)
    if want <= have:
        return
    anchor_col = rightmost_in(headers, anchor_members)
    if anchor_col is None:
        log.warning("No anchor column for family '%s' — skipping", primary)
        return
    for n in range(have + 1, want + 1):
        insert_at = anchor_col + 1
        ws.insert_cols(insert_at, len(insert_members))
        new_hdrs = [f"{m} {n}" for m in insert_members]
        for i, h in enumerate(new_hdrs):
            ws.cell(HEADER_ROW, insert_at + i).value = h
        # Shift headers dict right for moved columns
        shifted = {}
        for col, h in headers.items():
            shifted[col + len(insert_members) if col >= insert_at else col] = h
        for i, h in enumerate(new_hdrs):
            shifted[insert_at + i] = h
        headers.clear()
        headers.update(shifted)
        anchor_col += len(insert_members)   # keep anchor pointing at last inserted

def transplant(template_xlsx: str, edited_xlsx: str) -> None:
    out_dir = os.path.dirname(os.path.abspath(template_xlsx)) or "."
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=out_dir)
    os.close(fd)
    try:
        with (zipfile.ZipFile(template_xlsx, "r") as zt,
              zipfile.ZipFile(edited_xlsx, "r") as ze):
            t_names = {i.filename for i in zt.infolist()}
            e_names = {i.filename for i in ze.infolist()}
            replace_ws = [n for n in t_names
                          if n.startswith("xl/worksheets/") and n.endswith(".xml")
                          and n in e_names]
            replace_ss = "xl/sharedStrings.xml" in e_names
            with zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zn:
                for info in zt.infolist():
                    nm = info.filename
                    if nm in replace_ws:
                        continue
                    if replace_ss and nm == "xl/sharedStrings.xml":
                        continue
                    zn.writestr(info, zt.read(nm))
                for nm in replace_ws:
                    zn.writestr(nm, ze.read(nm))
                if replace_ss:
                    zn.writestr("xl/sharedStrings.xml", ze.read("xl/sharedStrings.xml"))
        os.replace(tmp, template_xlsx)
    finally:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass

def fix_shared_strings(xlsx_path: str) -> None:
    """Convert inlineStr cells to shared-string references for EPPlus compatibility."""
    out_dir = os.path.dirname(os.path.abspath(xlsx_path)) or "."
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=out_dir)
    os.close(fd)
    try:
        with zipfile.ZipFile(xlsx_path, "r") as z_in:
            names = set(z_in.namelist())
            sl: list[str] = []
            si: dict[str, int] = {}
            total = 0

            if "xl/sharedStrings.xml" in names:
                try:
                    root = ET.fromstring(z_in.read("xl/sharedStrings.xml"))
                    ns = root.tag.split("}")[0].strip("{")
                    for ssi in root.findall(f"{{{ns}}}si"):
                        s = "".join(t.text or "" for t in ssi.findall(f".//{{{ns}}}t"))
                        si[s] = len(sl); sl.append(s)
                except Exception:
                    sl, si = [], {}

            def _add(s: str) -> int:
                if s in si:
                    return si[s]
                idx = len(sl); si[s] = idx; sl.append(s); return idx

            cell_re = re.compile(
                rb'(<c\b[^>]*\bt="inlineStr"[^>]*>)(\s*<is>.*?</is>\s*)(</c>)', re.DOTALL)
            t_re    = re.compile(rb'<t(?:\s+[^>]*)?>(.*?)</t>', re.DOTALL)
            changed: dict[str, bytes] = {}

            for part in names:
                if not (part.startswith("xl/worksheets/") and part.endswith(".xml")):
                    continue
                xml = z_in.read(part)
                if b't="inlineStr"' not in xml:
                    continue

                def _sub(m: re.Match) -> bytes:
                    nonlocal total
                    texts = []
                    for tm in t_re.finditer(m.group(2)):
                        try:
                            texts.append(html.unescape(tm.group(1).decode("utf-8")))
                        except Exception:
                            texts.append(html.unescape(tm.group(1).decode("utf-8", "replace")))
                    idx = _add("".join(texts)); total += 1
                    # count MUST be a keyword: positional `count` is deprecated
                    # since CPython 3.13 and slated to become a TypeError. On
                    # 3.14 the positional form emitted a DeprecationWarning per
                    # rewritten cell — 6,923 per eidr-wikidata test run (T1,
                    # HANDOFF 2026-08-09).
                    return re.sub(rb't="inlineStr"', b't="s"', m.group(1), count=1) + \
                           f"<v>{idx}</v>".encode() + m.group(3)

                changed[part] = cell_re.sub(_sub, xml)

            if not changed:
                return

            ss_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
            root2 = ET.Element(f"{{{ss_ns}}}sst")
            root2.set("count", str(total)); root2.set("uniqueCount", str(len(sl)))
            for s in sl:
                ssi2 = ET.SubElement(root2, f"{{{ss_ns}}}si")
                t2 = ET.SubElement(ssi2, f"{{{ss_ns}}}t")
                if s.startswith(" ") or s.endswith(" ") or "\n" in s or "\t" in s:
                    t2.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
                t2.text = s
            new_ss = ET.tostring(root2, encoding="utf-8", xml_declaration=False)

            ct = z_in.read("[Content_Types].xml")
            if b"sharedStrings.xml" not in ct:
                ct = ct.replace(b"</Types>",
                    b'<Override PartName="/xl/sharedStrings.xml" '
                    b'ContentType="application/vnd.openxmlformats-officedocument'
                    b'.spreadsheetml.sharedStrings+xml"/></Types>')

            with zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as z_out:
                for info in z_in.infolist():
                    nm = info.filename
                    if nm in changed:
                        z_out.writestr(info, changed[nm])
                    elif nm == "xl/sharedStrings.xml":
                        z_out.writestr(info, new_ss)
                    elif nm == "[Content_Types].xml":
                        z_out.writestr(info, ct)
                    else:
                        z_out.writestr(info, z_in.read(nm))
                if "xl/sharedStrings.xml" not in names:
                    z_out.writestr("xl/sharedStrings.xml", new_ss)

        os.replace(tmp, xlsx_path)
    finally:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Reader half (2026-08-06)
# ---------------------------------------------------------------------------

@contextlib.contextmanager
def open_sheet(path: str, sheet_name: str, *,
               header_row: int = HEADER_ROW,
               data_start: int = DATA_START,
               stop: str | None = None,
               ) -> Iterator[tuple[dict[int, str],
                                   Iterator[tuple[int, dict[str, object]]]]]:
    """Open one BMR sheet for STREAMING reads: ``(headers, rows)``.

    Contributed by BMRtoAltID 2026-08-26 (`HANDOFF-bmr-io-streaming.md`),
    whose production sheets reach 60k rows by 500+ columns — roughly 39M
    dict entries if fully materialized, which is multi-GB before the
    values themselves. ``read_sheet``'s materializing shape is right for
    the chunked path it was seeded from and unusable for that one.

    A context manager, because the workbook handle has to outlive the
    header map: ``read_only`` mode holds an open file (and, for a zip64
    sheet, a temp file) that must be released deterministically. On
    Windows an unclosed handle keeps the .xlsx locked against the
    operator's own Excel, so the ``with`` block is not a formality — a
    bare generator would leave closing to garbage collection.

    ``headers`` is ``{1-based INTEGER column: header text}`` (trimmed,
    blanks skipped, NO truncation at header gaps) and is fully
    materialized before the block body runs, so a consumer can plan its
    column layout up front. The integer key is a CONTRACT — see
    ``_header_map``.

    ``rows`` yields ``(row_number, {header: value})`` LAZILY and is valid
    ONLY inside the ``with`` block. ``row_number`` is the sheet's own
    1-based absolute row, so it JUMPS over rows ``stop=None`` skipped:
    it cannot be reconstructed downstream, and a consumer reporting a
    defect to a human must be able to name the row they see in Excel.
    Blank cells are absent from the dict, so the blank sentinel is the
    consumer's choice (``row.get(h)`` vs ``row.get(h, "")``).

    ``stop`` selects the end-of-data rule (see module docstring) and is
    validated BEFORE the workbook is opened, so a typo fails without
    leaving a handle behind.
    """
    if stop not in (None, "blank_row", "blank_first_col"):
        raise ValueError(f"unknown stop rule: {stop!r}")
    # Lazy import keeps eidr_core.bmr_io importable without openpyxl for
    # consumers that only use the zip-surgery / layout / plan helpers.
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"{path}: expected sheet {sheet_name!r}, found "
                f"{wb.sheetnames!r}"
            )
        ws = wb[sheet_name]

        headers: dict[int, str] = {}
        for row_vals in ws.iter_rows(min_row=header_row, max_row=header_row,
                                     values_only=True):
            # Same policy as read_headers, shared via _header_map — only
            # the row ACCESS differs (streaming here, random-access there).
            headers = _header_map(row_vals)
            break

        n_cols = max(headers, default=0)
        headers_by_pos = [headers.get(i + 1) for i in range(n_cols)]

        def _rows() -> Iterator[tuple[int, dict[str, object]]]:
            for offset, row_vals in enumerate(
                    ws.iter_rows(min_row=data_start, values_only=True)):
                row_no = data_start + offset
                if stop == "blank_first_col":
                    first = row_vals[0] if row_vals else None
                    if first is None or (isinstance(first, str)
                                         and not first.strip()):
                        break
                row_dict: dict[str, object] = {}
                for i, val in enumerate(row_vals[:n_cols]):
                    if val is None or val == "":
                        continue
                    hdr = headers_by_pos[i]
                    if hdr:
                        row_dict[hdr] = val
                if not row_dict:
                    if stop == "blank_row":
                        break
                    continue    # stop=None: blank row is skippable padding
                yield row_no, row_dict

        yield headers, _rows()
    finally:
        wb.close()


def read_sheet(path: str, sheet_name: str, *,
               header_row: int = HEADER_ROW,
               data_start: int = DATA_START,
               stop: str | None = None,
               ) -> tuple[dict[int, str], list[dict[str, object]]]:
    """Read one BMR sheet: ``(headers, rows)``, rows fully materialized.

    Identical semantics to :func:`open_sheet` — a thin ``list()`` over
    it, so the header policy and all three stop rules keep exactly ONE
    implementation and the two entry points cannot drift — minus the row
    numbers, which predate that call and no existing caller asked for
    (adding them here would be a breaking change).

    ``headers`` is ``{1-based INTEGER column: header text}``; the integer
    key is a CONTRACT consumers sort on to recover sheet order (see
    ``_header_map``). ``rows`` is one ``{header: value}`` dict per data
    row, blank cells absent.

    Prefer ``open_sheet`` when the sheet may be large enough that holding
    every row costs more than the convenience is worth — a 60k-row,
    500-column BMR sheet is multi-GB here.
    """
    with open_sheet(path, sheet_name, header_row=header_row,
                    data_start=data_start, stop=stop) as (headers, rows):
        return headers, [row for _row_no, row in rows]


def family_layout(header_names: Iterable[str], members: Sequence[str],
                  ) -> dict[int, dict[str, str]]:
    """Map a repeating family's headers to ``{group_index: {member: header}}``.

    Group indices are the sheet's OWN numbering, gaps preserved (sparse) —
    callers densify only after pairing companion columns, never before.
    A bare member name is group 1; ``"Member N"`` (one-or-more spaces,
    digits) is group N, and beats the bare form for group 1.
    """
    pats = {m: re.compile(re.escape(m) + r"\s+(\d+)$") for m in members}
    out: dict[int, dict[str, str]] = {}
    for h in header_names:
        for m in members:
            if h == m:
                out.setdefault(1, {}).setdefault(m, h)
            else:
                mt = pats[m].fullmatch(h)
                if mt:
                    out.setdefault(int(mt.group(1)), {})[m] = h
    return out


class RepeatPlan:
    """Union-max repeat-group counts across a record set.

    ``bump`` records the widest instance seen per family; ``finalize``
    applies the construction-time minimums and floors every recorded
    family at 1 (a family that appeared at all gets at least one column
    group). ``get`` never returns less than 1 — header layouts always
    carry one group even for families empty across the whole set.
    """

    def __init__(self, minimums: Mapping[str, int] | None = None):
        self.counts: dict[str, int] = {}
        self.minimums: dict[str, int] = dict(minimums or {})

    def bump(self, family: str, n: int | None) -> None:
        if n is None:
            return
        self.counts[family] = max(self.counts.get(family, 0), int(n))

    def finalize(self) -> None:
        for fam, minv in self.minimums.items():
            self.counts[fam] = max(self.counts.get(fam, 0), minv)
        for fam in list(self.counts.keys()):
            self.counts[fam] = max(self.counts[fam], 1)

    def get(self, family: str, default: int = 1) -> int:
        v = self.counts.get(family)
        return max(v if v is not None else default, 1)


def pad_groups(items: Sequence, count: int, width: int,
               fill: str = "") -> list:
    """Flatten one family's values to exactly ``count`` groups of
    ``width`` cells, padding missing groups with ``fill``.

    Each item is a tuple/list of ``width`` cells, or a scalar when
    ``width == 1`` (e.g. a plain country-code list).
    """
    out: list = []
    for idx in range(count):
        if idx < len(items):
            item = items[idx]
            if isinstance(item, (tuple, list)):
                if len(item) != width:
                    raise ValueError(
                        f"group {idx + 1} has {len(item)} cells, expected {width}"
                    )
                out.extend(item)
            elif width == 1:
                out.append(item)
            else:
                raise ValueError(
                    f"scalar group value with width={width}; pass tuples"
                )
        else:
            out.extend([fill] * width)
    return out
